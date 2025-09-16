import csv
from django.http import HttpResponse
import openpyxl
from abc import ABC, abstractmethod
from typing import List, Dict,Any

class BaseExporter(ABC):
    """абстрактный базовый класс для экспорта в разные форматы"""
    @abstractmethod
    def export(self,contacts: List[Dict[str,Any]]) -> HttpResponse:
        pass
    def _prepare_contact_data(self, contact: Dict[str,Any]) -> Dict[str,str]:
        """подготовка данных контакта к экспорту"""
        return {
            'имя': contact.get('NAME', ''),
            'фамилия': contact.get('LAST_NAME', ''),
            'номер телефона': contact.get('PHONE', ''),
            'почта': contact.get('EMAIL', ''),
            'компания': contact.get('COMPANY', ''),
        }

class  CSVExporter(BaseExporter):
    """Экспортирует в .csv формат"""
    def export(self,contacts: List[Dict[str,Any]]) -> HttpResponse:
        response = HttpResponse(content_type='text/csv; charset=utf-16')
        response['Content-Disposition'] = 'attachment; filename="contact_export.csv"'

        fieldnames = ['имя', 'фамилия', 'номер телефона', 'почта', 'компания']
        writer = csv.DictWriter(response, fieldnames=fieldnames)
        writer.writeheader()

        for contact in contacts:
            contact_data = self._prepare_contact_data(contact)
            writer.writerow(contact_data)
        return response

class ExcelExporter(BaseExporter):
    """Экспорт в .xlsx формат"""

    def export(self,contacts: List[Dict[str,Any]]) -> HttpResponse:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Контакты'

        # заголовки
        headers = ['имя', 'фамилия', 'номер телефона', 'почта', 'компания']
        for col_num,  header in enumerate(headers,1):
            ws.cell(row=1, column=col_num,value=header)

        # данные
        for row_num, contact in enumerate(contacts,2):
            contact_data = self._prepare_contact_data(contact)
            for header_index, header_title in enumerate(headers,1):
                ws.cell(row=row_num, column=header_index,value=contact_data[header_title] or '')

        # автоподбор ширрины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as e:
                    raise ValueError(f'ошибка при обработке строки: {str(e)}')
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="contact_export.xlsx"'

        wb.save(response)
        return response

class ExporterFactory:
    """фабрика для создания экспортеров"""
    _exporters = {
        'csv': CSVExporter,
        'xlsx': ExcelExporter,
    }

    @classmethod
    def get_exporter(cls, format_type: str) -> BaseExporter:
        exporter_class = cls._exporters.get(format_type)
        if not exporter_class:
            raise ValueError(f'Неиспользуемый формат экспорта: {format_type}')
        return exporter_class()

    @classmethod
    def register_exporter(cls, format_type:str, exporter_class):
        """метод для регистрации новых экспортеров"""
        cls._exporters[format_type] = exporter_class