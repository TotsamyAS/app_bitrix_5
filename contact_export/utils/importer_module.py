import csv
import os
from random import sample

import openpyxl
from abc import ABC, abstractmethod
from typing import List, Dict, Any, Optional
from io import TextIOWrapper
from django.core.files.uploadedfile import UploadedFile

class BaseImporter(ABC):
    """абстрактный базовый класс для импортёров"""

    @abstractmethod
    def import_file(self, file: UploadedFile) -> List[Dict[str, Any]]:
        pass
    def _validate_row(self, row:  Dict[str, Any]) -> bool:
        """валидация строковых данных"""
        # минимальная валидация будет в том, что необходимы хотя бы имя + фамилия
        name = row.get('имя', '').strip()
        last_name = row.get('фамилия','').strip()
        return bool(name or last_name)

    def _normalize_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        """нормализация строковых данных"""
        normalized = {
            'NAME': row.get('имя','').strip(),
            'LAST_NAME': row.get('фамилия', '').strip(),
            'PHONE': row.get('номер телефона', '').strip(),
            'EMAIL': row.get('почта', '').strip(),
            'COMPANY_NAME': row.get('компания', '').strip(),
        }

        # очистка от пробелов
        for key in normalized:
            if normalized[key] is not None:
                normalized[key] = normalized[key].strip()
            else:
                normalized[key] = ''

            # очистка от пустых значений
        for key in normalized:
            if normalized[key] == '':
                normalized[key] = None

        return normalized

class CSVImporter(BaseImporter):
    """импортирует из .csv-формата"""
    def import_file(self, file: UploadedFile) -> List[Dict[str, Any]]:
        contacts = []
        # декодирование файла с правильной кодировкой
        try:
            # сброс файла на начало
            file.file.seek(0)
            text_content = file.file.read()

            decoded_content = text_content.decode('utf-16')
            if decoded_content is None:
                raise ValueError("не удалось декодировать файл с поддерживаемыми кодировками")
            #убрать BOM если есть
            if decoded_content.startswith('\ufeff'):
                decoded_content = decoded_content[1:]
                # print('>> BOM удалён')
            lines = decoded_content.splitlines()
            if not lines:
                raise ValueError("файл пуст")
            # print(f">> получено строк: {len(lines)}")
            # print(f">> первые 2 строки: {lines[:2]}")

            cleaned_lines = []
            for line in lines:
                # Убираем BOM из каждой строки
                if '\ufeff' in line:
                    line = line.replace('\ufeff', '')
                    line = line.replace('"', '')

                cleaned_lines.append(line)

            # print(f">> очищенные строки: {cleaned_lines[:2]}")

            # определение разделителя
            first_line = cleaned_lines[0]
            possible_delimiters = [',',';','\t']
            delimiter = ','

            for delimiter_variant in possible_delimiters:
                if delimiter_variant in first_line:
                    delimiter = delimiter_variant
                    break
            # print(f">> определен разделитель: '{delimiter}'")

            reader = csv.DictReader(cleaned_lines, delimiter=delimiter)
            # print(f'>> определены заголовки: {reader.fieldnames}')

            possible_abbreviations = ['ООО', 'ОАО', 'ИП', 'ЗАО', 'ПАО', 'НПАО', 'ГУП', 'МУП']

            for row_num, row in enumerate(reader, 1):
                try:
                    # print(f">> сырая строка {row_num}: {dict(row)}")

                    # проверяем, что значения не None
                    cleaned_row = {}
                    for key, value in row.items():
                        if value is None:
                            cleaned_row[key] = ''
                        else:
                            cleaned_row[key] = value

                    if self._validate_row(row):
                        normalized_row = self._normalize_row(row)
                        if normalized_row['COMPANY_NAME'].split(' ')[0] in possible_abbreviations:
                            company_name_splitted = normalized_row['COMPANY_NAME'].split(' ')
                            normalized_row['COMPANY_NAME'] = (
                                    company_name_splitted[0]
                                    + ' "'
                                    + ' '.join(company_name_splitted[1:])
                                    +'"')
                        contacts.append(normalized_row)
                        # print(f">> строка {row_num} прошла валидацию: {normalized_row}")
                    else:
                        # print(f">> строка {row_num} не прошла валидацию: {row}")
                        pass
                except Exception as e:
                    print(f">> Ошибка в строке {row_num}: {e}")
                    continue

            print(f">> успешно обработано контактов: {len(contacts)}")
            return contacts

        except Exception as e:
            print(f">> критическая ошибка при обработке CSV: {str(e)}")
            raise ValueError(f'ошибка при обработке CSV файла: {str(e)}')


class XLSXImporter(BaseImporter):
    """импортирует из xlsx файлов"""

    def import_file(self, file: UploadedFile) -> List[Dict[str, Any]]:
        contacts = []

        wb = openpyxl.load_workbook(filename=file.file, read_only=True)
        ws = wb.active

        # считывание заголовков
        headers = []
        for cell in ws[1]:
            headers.append(cell.value.strip().lower() if cell else '')

        # считываем данные
        for row_num in range(2, ws.max_row + 1):
            row_data = {}
            for col_num, header in enumerate(headers, 1):
                cell_value = ws.cell(row=row_num, column=col_num).value
                row_data[header] = str(cell_value) if cell_value is not None else ''

            try:
                if self._validate_row(row_data):
                    normalized_row = self._normalize_row(row_data)
                    contacts.append(normalized_row)
            except Exception as e:
                print(f'Ошибка при импорте из .xslx файла {e}')
                raise ValueError(f'Ошибка при импорте из .xslx файла {e}')
                return []

        return contacts


class ImporterFactory:
    """фабрика для создания новых импортов"""
    _importers = {
        'csv': CSVImporter,
        'xlsx': XLSXImporter,
    }

    @classmethod
    def get_importer(cls, format_type: str) -> BaseImporter:
        importer_class = cls._importers.get(format_type.lower())
        if not importer_class:
            raise ValueError(f'неподдерживаемый формат импорта {format_type}')
        return importer_class()

    @classmethod
    def register_importer(cls, format_type: str, importer_class):
        """регистрация новых импортеров"""
        cls._importers[format_type.lower()] = importer_class
        
    
    
# утилиты для работы с импортом
def detect_file_format(filename: str) -> str:
    """опрределение формата файла по расширению"""
    extension = filename.split('.')[-1].lower()
    if extension == 'csv':
        return 'csv'
    elif extension in ['xlsx', 'xls']:
        return 'xlsx'
    else:
        print(f'неподдерживаемый формат файла {extension}')
        raise ValueError(f'неподдерживаемый формат файла {extension}')


def process_imported_file(file: UploadedFile) -> List[Dict[str, Any]]:
    """основная функция обработки загруженного файла"""
    try:
        format_type = detect_file_format(file.name)
        importer = ImporterFactory.get_importer(format_type)
        contacts= importer.import_file(file)
        return contacts
    except Exception as e:
        ValueError(f'Ошибка при обработке файла {file.name}: {str(e)}')